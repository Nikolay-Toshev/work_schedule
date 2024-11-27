from openpyxl import Workbook
from openpyxl.styles import Alignment
from db_queries import get_employees, add_month, get_working_hours_by_day_and_week_schedule


def add_month_title(month, year):
    ws.merge_cells("A4:D5")
    ws["A4"] = f"месец {month} {year}"
    ws["A4"].alignment = alignment


def vacation_days():
    ws['F3'] = 'Отпуск:'


def add_column_title():
    ws["A8"] = "дата"
    ws["A8"].alignment = alignment
    ws.column_dimensions['A'].width = 12

    ws["B8"] = "служител"
    ws["B8"].alignment = alignment
    ws.column_dimensions['B'].width = 10

    ws["C8"] = "график\nот час до час"
    ws["C8"].alignment = alignment
    ws.column_dimensions['C'].width = 13

    ws["D8"] = "отработени\nчасове"
    ws["D8"].alignment = alignment
    ws.column_dimensions['D'].width = 12

    ws.column_dimensions['E'].width = 4

    ws["F8"] = "работни\nдни"
    ws["F8"].alignment = alignment

    ws.column_dimensions['G'].width = 3

    ws["H8"] = "часове"
    ws["H8"].alignment = alignment

    ws.row_dimensions[8].height = 30


def add_days_to_table(i, n, work_month, num_of_employees):
    ws.merge_cells(f"A{n}:A{n + (num_of_employees - 1)}")
    ws[f"A{n}"] = f"{work_month[i][0]}\n{work_month[i][1]}"
    ws[f"A{n}"].alignment = Alignment(horizontal='center', vertical='top', wrapText=True)


def add_employees_and_hours(n, day, week_schedule, num_of_employees):
    workhours = get_working_hours_by_day_and_week_schedule(day, week_schedule)

    for i in range(num_of_employees):
        ws[f"B{(n + i)}"] = workhours[i][0]

        ws[f"C{(n + i)}"] = workhours[i][1]
        ws[f"C{(n + i)}"].alignment = Alignment(horizontal='center')

        ws[f"D{(n + i)}"] = workhours[i][2]
        ws[f"D{(n + i)}"].alignment = Alignment(horizontal='center')

def calculate_hours_per_day(n, num_of_employees):
    for i in range(num_of_employees):
        ws[f"E{(n + i)}"] = f"=D{(n + i)} - 8"


def create_hours_per_month_formula(n, num_of_employees):
    for i in range(num_of_employees):
        ws[f"H{(9 + i)}"] = f'=SUM(D{(n + i)}'


def add_to_hours_per_month_formula(n, num_of_employees):
    if n > 9:
        for i in range(num_of_employees):
            ws[f"H{(9 + i)}"].value += f',D{(n + i)}'


def calculate_extra_hours_per_month(num_of_employees):
    for i in range(num_of_employees):
        ws[f"F{(9 + i)}"] = f'=H{(9 + i)}-I8'


def workdays_per_month(workdays):
    ws['G8'] = workdays
    ws['G8'].alignment = Alignment(horizontal='center', vertical='top')


def workhours_per_month():
    ws['I8'] = '=G8*8'
    ws['I8'].alignment = Alignment(horizontal='center', vertical='top')


def create_table(year, month, workdays, option_menu_weeks):
    work_month = add_month(month, year)
    num_of_employees = len(get_employees())

    n = 9
    add_month_title(month, year)
    vacation_days()
    add_column_title()
    workdays_per_month(workdays)
    workhours_per_month()
    create_hours_per_month_formula(n, num_of_employees)
    week_schedules = [(empl_id, week.cget('text')) for empl_id, week in option_menu_weeks.items()]
    w = 0
    for i in range(len(work_month)):
        week_schedule = week_schedules[w][1]
        if work_month[i][1] == 'Неделя':
            w += 1
        add_days_to_table(i, n, work_month, num_of_employees)
        add_employees_and_hours(n, work_month[i][1], week_schedule, num_of_employees)
        calculate_hours_per_day(n, num_of_employees)
        add_to_hours_per_month_formula(n, num_of_employees)
        n += num_of_employees
    calculate_extra_hours_per_month(num_of_employees)


wb: Workbook = Workbook()

ws = wb.active

alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)


if __name__ == '__main__':
    add_days_to_table(2023, 8)
    wb.save("test.xlsx")
